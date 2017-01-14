import csv
import glob
import openpyxl

# create excel output file
output_workbook = openpyxl.workbook()
data_sheet = output_workbook.create_sheet("data")
current_row = current_column = 1

# write header to excel output file
column_labels = [
    "participant_id",
    "trial_id",
    "trial_start",
    "trial_type",
    "trial_stim",
    "b_char_shown_time",
    "b_char_resp_yn",
    "b_char_resp_time"
]
for column_label in column_labels:
    data_sheet.cell(row=current_row, column=current_column, value=column_label)
    current_column += 1
current_column = 1
current_row = 2

# define global variables for trial data
trial_id = trial_start = trial_type = trial_stim = none
b_char_list = []
condition_set = set()
participant_set = set()

# loop over data files in pilotdata folder
for file in glob.glob("../pilotdata/*"):
    with open(file) as f:
        # create a csv reader for each file
        csv_reader = csv.reader(f, delimiter=",")
        # loop over the rows in each file
        for row in csv_reader:
            # extract information from trialstart event
            if row[3] == "trialstart":
                trial_id = row[1]
                trial_start = int(row[2])
                trial_type = row[4]
                trial_stim = row[5]
                condition_set.add(trial_type)  # add trial_type to condition_set for use in analysis
                participant_set.add(row[0])  # add participant_id to participant_set for use in analysis
            # extract information from charshow event
            if row[3] == "charshow":
                # if you find a "b", write new row to spreadsheet
                if row[6] == "b":
                    b_char_list = [
                        row[0],  # participant_id
                        trial_id,
                        trial_start,
                        trial_type,
                        trial_stim,
                        int(row[2]),  # b_char_show_time
                        "",  # b_char_resp_yn, leave empty
                        "",  # b_char_resp_time, leave empty
                    ]
                    for i in b_char_list:
                        data_sheet.cell(row=current_row, column=current_column, value=i)
                        current_column += 1
                    current_column = 1
                    current_row += 1
            # extract information from screentap event
            if row[3] == "screentap":
                # only register first tap after "b" is shown
                try:
                    if b_char_list[6] == "":
                        b_char_list[6] = "y"  # b_char_resp_yn
                    if b_char_list[7] == "":
                        b_char_list[7] = int(row[2]) - b_char_list[5]  # b_char_resp_time
                    # write updated b_char_list on previous row
                    for i in b_char_list:
                        data_sheet.cell(row=current_row - 1, column=current_column, value=i)
                        current_column += 1
                    current_column = 1
                except indexerror:
                    print("current row: {}\ncurrent b_char_list: {}".format(row, b_char_list))

# write "n" in empty cells in b_char_resp_yn column
for cell in data_sheet.columns[6]:
    if cell.value not in ["b_char_resp_yn", "y"]:
        cell.value = "n"

# calculate averages per condition type
analysis_sheet = output_workbook.create_sheet("analysis")
analysis_sheet.cell(row=1, column=1, value="overall average rts")
current_row = 2
current_column = 1
for condition in condition_set:
    analysis_sheet.cell(row=current_row, column=current_column, value=condition)
    current_row += 1
    condition_total = condition_count = 0  # resp. sum and number of rts in condition
    # loop through data sheet to find entries with the current condition type
    for cell in data_sheet.columns[3]:
        if cell.value == condition:
            condition_count += 1
            # add rt to condition_total
            try:
                condition_total += int(data_sheet.cell(row=cell.row, column=8).value)
            except valueerror:
                condition_total += 0
    # get average rt for condition
    try:
        condition_average = float(condition_total / condition_count)
    except zerodivisionerror:
        print("error: " + condition)
        condition_average = 0
    analysis_sheet.cell(row=current_row, column=current_column, value=condition_average)
    current_row -= 1
    current_column += 1

# calculate averages per condition type per participant
analysis_sheet.cell(row=5, column=1, value="average rts per participant")
analysis_sheet.cell(row=6, column=1, value="participant")
current_row = 6
current_column = 1
for participant in participant_set:
    analysis_sheet.cell(row=current_row+1, column=current_column, value=participant)
    current_column += 1
    for condition in condition_set:
        analysis_sheet.cell(row=6, column=current_column, value=condition)
        current_row += 1
        condition_total = condition_count = 0
        # loop through data sheet to find entries of the current condition and participant
        for cell in data_sheet.columns[3]:
            if cell.value == condition and data_sheet.cell(row=cell.row, column=1).value == participant:
                condition_count += 1
                # add rt to condition_total
                try:
                    condition_total += int(data_sheet.cell(row=cell.row, column=8).value)
                except valueerror:
                    condition_total += 0
        # get average rt for condition
        try:
            condition_average = float(condition_total / condition_count)
        except zerodivisionerror:
            print("error: " + condition)
            condition_average = 0
        analysis_sheet.cell(row=current_row, column=current_column, value=condition_average)
        current_row -= 1
        current_column += 1
    current_row += 1
    current_column = 1

# remove empty sheet 'sheet' and save output file
output_workbook.remove_sheet(output_workbook.worksheets[0])
output_workbook.save("../analyzed_data.xlsx")

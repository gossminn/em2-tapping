import csv
import glob
import re
import statistics

import openpyxl

from session import Session


class TappingData:
    def __init__(self, pathname, include_practice=False):
        # List of data sessions
        self.sessions = []

        # Define data file directory
        files = glob.glob("{}/*".format(pathname))

        # Loop through data files
        for file in files:
            if not include_practice:
                if re.search("practice", file): 
                    continue

            # Get participant id and session type from file
            with open(file) as f1:
                # Get contents of csv file
                csv_reader = csv.reader(f1, delimiter=",")
                csv_contents = list(csv_reader)[1:]

                # Extract information from first row
                csv_start = csv_contents[0]
                subject_id = csv_start[0]
                session_type = csv_start[4] if csv_start[4] == "practice" else "test"

            # Create session object and append to list
            session = Session(subject_id, session_type, csv_contents)
            self.sessions.append(session)

        # Statistics
        self.rt_per_condition = self.get_condition_rts()
        self.rts_per_condition = self.get_condition_rtlist()
        self.errs_per_condition = self.get_condition_errs()
        self.drts_per_condition = self.get_condition_drts()
        self.rrts_per_condition = self.get_condition_rrts()

    def write_sessions_xlsx(self):
        # New empty workbook
        wb = openpyxl.Workbook()

        # Log every participant's responses on a separate sheet
        for session in self.sessions:
            # Create a new sheet
            sheet_name = "Subject_{}".format(session.subj_id)
            wb.create_sheet(title=sheet_name)
            current_sheet = wb.get_sheet_by_name(sheet_name)

            # Set initial row and column values
            current_col = 1
            current_row = 1

            # First column labels (+ header)
            current_cell = current_sheet.cell(row=current_row, column=current_col)
            current_cell.value = "Resp_ID"

            for i in range(1, 35):
                current_row += 1
                current_cell = current_sheet.cell(row=current_row, column=current_col)
                current_cell.value = i

            current_row = 1

            # Trials (+ column headers)
            for i in range(1, 1 + len(session.trials)):
                # Trial_1, Trial_2, ...
                current_col += 1
                current_cell = current_sheet.cell(row=current_row, column=current_col)
                current_cell.value = "Trial_{}".format(i)

            current_col = 1

            for trial in session.trials:
                current_col += 1
                current_row = 2
                for response in trial.responses:
                    # Write RTs
                    current_cell = current_sheet.cell(row=current_row, column=current_col)
                    current_cell.value = response.delta_time if not response.delta_time == 0 else ""
                    current_row += 1

        # Save the workbook
        wb.save("../ResponseData.xlsx")

    def write_session_rrts(self):
        # New workbook and sheet
        wb = openpyxl.Workbook()
        sheet = wb.active

        # Column, row counters
        current_row = 1
        current_col = 1

        # Write mean rrts per session
        for condition in self.rrts_per_condition:
            # Column header
            sheet.cell(row=current_row, column=current_col).value = condition

            # Mean rrts for this condition
            for i in self.rrts_per_condition[condition]:
                current_row += 1
                sheet.cell(row=current_row, column=current_col).value = i if i != 0 else ""

            # Next column, reset row
            current_col += 1
            current_row = 1

        # Save workbook
        wb.save("../ConditionRrts.xlsx")

    def get_condition_rts(self):
        # Get condition-rts per session
        rts_per_session = [s.rt_per_condition for s in self.sessions]

        # Empty output dictionary
        rts = {}

        # Get keys
        for s in rts_per_session:
            for key in s:
                rts[key] = []

        # Get list of session mean rts per condition
        for key in rts:
            for s in rts_per_session:
                rts[key].append(s[key])

        # Get overall mean rts per condition
        for key in rts:
            rts[key] = statistics.mean(rts[key])

        return rts

    def get_condition_rtlist(self):
        # Get condition-rts per session
        rts_per_session = [s.rt_per_condition for s in self.sessions]

        # Empty output dictionary
        rts = {}

        # Get keys
        for s in rts_per_session:
            for key in s:
                rts[key] = []

        # Get list of session mean rts per condition
        for key in rts:
            for s in rts_per_session:
                rts[key].append(s[key])

        return rts

    def get_condition_errs(self):
        # Get condition-errs per session
        errs_per_session = [s.errors_per_condition for s in self.sessions]

        # Emtpy output dictionary
        errs = {}

        # Get keys
        for s in errs_per_session:
            for key in s:
                errs[key] = []

        # Get list of session mean errs per condition
        for key in errs:
            for s in errs_per_session:
                errs[key].append(s[key])

        # Get overall mean errs per condition
        for key in errs:
            errs[key] = statistics.mean(errs[key])

        return errs

    def get_condition_drts(self):
        # Condition-drts per session
        drts_per_session = [s.delta_rt_per_condition for s in self.sessions]

        # Empty output dictionary
        deltas = {}

        # Get keys
        for s in drts_per_session:
            for key in s:
                deltas[key] = []

        # Get list of session mean drts per condition
        for key in deltas:
            for s in drts_per_session:
                deltas[key].append(s[key])

        # Get overall mean drts per condition
        for key in deltas:
            deltas[key] = statistics.mean(deltas[key])

        return deltas

    def get_condition_rrts(self):
        # Condition-rrts per session
        rrts_per_session = [s.resp_rts_per_condition for s in self.sessions]

        # Empty output dictionary
        rrts = {}

        # Get keys
        for s in rrts_per_session:
            for key in s:
                rrts[key] = []

        # Get mean rrts list per condition
        for key in rrts:
            # Count up to 35 responses
            for i in range(35):
                # List of mean RTs for this resp_id
                rts = []
                for s in rrts_per_session:
                    rts.append(s[key][i])

                # Average RT for this resp_id
                rrts[key].append(statistics.mean(rts))
        return rrts
